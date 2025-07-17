import os
import tempfile
import csv

import pyodbc
from flask import (
    Flask,
    request,
    render_template_string,
    send_file,
    after_this_request,
    redirect,
    url_for,
)
from openpyxl import Workbook, load_workbook
from pycomm3 import LogixDriver


def export_instruments_to_excel(mdb_path: str, xlsx_path: str) -> int:
    """Export the Instruments table to an Excel workbook.

    Only rows with Type='IO' are exported. The columns Tag, FullDescription,
    EGULow, EGUHigh, RawLow, RawHigh and a set of alarm/warning columns are
    written. The instruments are grouped into DigitalInput, DigitalOutput,
    AnalogInput and AnalogOutput sheets. Returns the number of rows exported.
    """

    conn_str = (
        r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
        f'DBQ={mdb_path};'
    )
    conn = pyodbc.connect(conn_str, autocommit=True)
    cursor = conn.cursor()

    # Verify the Instruments table exists
    table_names = [row.table_name for row in cursor.tables(tableType='TABLE')]
    if 'Instruments' not in table_names:
        raise ValueError('Instruments table not found')

    query = (
        "SELECT ID, Tag, FullDescription, EGULow, EGUHigh, RawLow, RawHigh, "
        "HALM_EN, HALM_SP, HALM_DB, HALM_DLY, "
        "HWARN_EN, HWARN_SP, HWARN_DB, HWARN_DLY, "
        "LALM_EN, LALM_SP, LALM_DB, LALM_DLY, "
        "LWARN_EN, LWARN_SP, LWARN_DB, LWARN_DLY, "
        "DigitalInput, DigitalOutput, AnalogInput, AnalogOutput "
        "FROM Instruments WHERE Type='IO' AND Tag <> '' AND Tag IS NOT NULL"
    )
    cursor.execute(query)
    rows = cursor.fetchall()

    header = [
        'ID',
        'Tag',
        'FullDescription',
        'EGULow',
        'EGUHigh',
        'RawLow',
        'RawHigh',
        'HALM_EN',
        'HALM_SP',
        'HALM_DB',
        'HALM_DLY',
        'HWARN_EN',
        'HWARN_SP',
        'HWARN_DB',
        'HWARN_DLY',
        'LALM_EN',
        'LALM_SP',
        'LALM_DB',
        'LALM_DLY',
        'LWARN_EN',
        'LWARN_SP',
        'LWARN_DB',
        'LWARN_DLY',
    ]

    categories = {
        'DigitalInput': [],
        'DigitalOutput': [],
        'AnalogInput': [],
        'AnalogOutput': [],
    }

    for row in rows:
        if getattr(row, 'DigitalInput', False):
            categories['DigitalInput'].append(row)
        elif getattr(row, 'DigitalOutput', False):
            categories['DigitalOutput'].append(row)
        elif getattr(row, 'AnalogInput', False):
            categories['AnalogInput'].append(row)
        elif getattr(row, 'AnalogOutput', False):
            categories['AnalogOutput'].append(row)

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    for name, rows_list in categories.items():
        ws = wb.create_sheet(name)
        ws.append(header)
        for row in rows_list:
            ws.append(row[:len(header)])

    wb.save(xlsx_path)

    conn.close()
    return sum(len(v) for v in categories.values())


def update_instruments_from_excel(mdb_path: str, excel_path: str) -> int:
    """Update the Instruments table using data from an exported Excel workbook.

    The Excel file must contain sheets named DigitalInput, DigitalOutput,
    AnalogInput and AnalogOutput with the same columns as produced by
    ``export_instruments_to_excel``. Only rows with matching ``ID`` and ``Tag``
    are updated. The function returns the number of rows that were modified.
    """

    expected_header = [
        'ID',
        'Tag',
        'FullDescription',
        'EGULow',
        'EGUHigh',
        'RawLow',
        'RawHigh',
        'HALM_EN',
        'HALM_SP',
        'HALM_DB',
        'HALM_DLY',
        'HWARN_EN',
        'HWARN_SP',
        'HWARN_DB',
        'HWARN_DLY',
        'LALM_EN',
        'LALM_SP',
        'LALM_DB',
        'LALM_DLY',
        'LWARN_EN',
        'LWARN_SP',
        'LWARN_DB',
        'LWARN_DLY',
    ]

    wb = load_workbook(excel_path, data_only=True)

    conn_str = (
        r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
        f'DBQ={mdb_path};'
    )
    conn = pyodbc.connect(conn_str, autocommit=False)
    cursor = conn.cursor()

    table_names = [row.table_name for row in cursor.tables(tableType='TABLE')]
    if 'Instruments' not in table_names:
        conn.close()
        raise ValueError('Instruments table not found')

    total_updates = 0

    try:
        for sheet_name in ['DigitalInput', 'DigitalOutput', 'AnalogInput', 'AnalogOutput']:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f'Sheet {sheet_name} missing from Excel file')
            ws = wb[sheet_name]
            header = list(next(ws.iter_rows(max_row=1, values_only=True)))
            if header != expected_header:
                raise ValueError(f'Invalid header in sheet {sheet_name}')

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if all(cell is None for cell in row):
                    continue
                data = dict(zip(expected_header, row))

                id_val = data['ID']
                tag_val = data['Tag']
                cursor.execute(
                    'SELECT ' + ', '.join(expected_header) + ' FROM Instruments WHERE ID=? AND Tag=?',
                    (id_val, tag_val)
                )
                db_row = cursor.fetchone()
                if not db_row:
                    raise ValueError(f'Row {row_idx} in sheet {sheet_name} has unknown ID/Tag')

                updates = []
                params = []
                for col in expected_header:
                    if col in ('ID', 'Tag'):
                        continue
                    excel_val = data[col]
                    db_val = getattr(db_row, col)
                    if excel_val != db_val:
                        updates.append(f'{col}=?')
                        params.append(excel_val)

                if updates:
                    params.extend([id_val, tag_val])
                    cursor.execute(
                        f"UPDATE Instruments SET {', '.join(updates)} WHERE ID=? AND Tag=?",
                        params
                    )
                    total_updates += 1

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return total_updates

HOME_TEMPLATE = """
<!doctype html>
<html lang='en'>
<head>
  <meta charset='utf-8'>
  <title>PLCPoke 2.0</title>
  <style>
    body { font-family: Arial, sans-serif; background:#f2f2f2; margin:40px; }
    .container { max-width: 600px; margin:auto; background:#fff; padding:40px; border-radius:8px; box-shadow:0 2px 4px rgba(0,0,0,0.1); text-align:center; }
    a.button { display:inline-block; padding:12px 24px; margin:10px; text-decoration:none; background:#007BFF; color:#fff; border-radius:4px; }
  </style>
</head>
<body>
  <div class="container">
    <h1>PLCPoke 2.0</h1>
    <a class="button" href="{{ url_for('export_page') }}">Export Instruments</a>
    <a class="button" href="{{ url_for('import_excel') }}">Update From Excel</a>
    <a class="button" href="{{ url_for('plc_page') }}">Read PLC Info</a>
  </div>
</body>
</html>
"""

EXPORT_TEMPLATE = """
<!doctype html>
<html lang='en'>
<head>
  <meta charset='utf-8'>
  <title>Export Instruments</title>
  <style>
    body { font-family: Arial, sans-serif; background:#f2f2f2; margin:40px; }
    .container { max-width:600px; margin:auto; background:#fff; padding:40px; border-radius:8px; box-shadow:0 2px 4px rgba(0,0,0,0.1); }
    #file-name { margin-left:10px; }
    button { padding:8px 16px; margin-top:10px; }
  </style>
</head>
<body>
  <div class="container">
    <h1>Export Instruments</h1>
    <form method="post" enctype="multipart/form-data">
      <input id="mdb-file" type="file" name="file" accept=".mdb" style="display:none" />
      <button type="button" onclick="document.getElementById('mdb-file').click()">Select File</button>
      <span id="file-name"></span>
      <button type="submit">Upload</button>
    </form>
    <script>
      const input = document.getElementById('mdb-file');
      const fileName = document.getElementById('file-name');
      input.addEventListener('change', () => {
        const file = input.files[0];
        fileName.textContent = file ? file.name : '';
      });
    </script>
    {% if message %}
    <p>{{ message }}</p>
    {% endif %}
    {% if excel_available %}
    <p><a href="{{ url_for('download_excel') }}">Download Instruments Excel</a></p>
    {% endif %}
    <p><a href="{{ url_for('home') }}">Back to Home</a></p>
  </div>
</body>
</html>
"""

IMPORT_TEMPLATE = """
<!doctype html>
<html lang='en'>
<head>
  <meta charset='utf-8'>
  <title>Update MDB</title>
  <style>
    body { font-family: Arial, sans-serif; background:#f2f2f2; margin:40px; }
    .container { max-width:600px; margin:auto; background:#fff; padding:40px; border-radius:8px; box-shadow:0 2px 4px rgba(0,0,0,0.1); }
    .file-name { margin-left:10px; }
    button { padding:8px 16px; margin-top:10px; }
  </style>
</head>
<body>
  <div class="container">
    <h1>Update MDB From Excel</h1>
    <form method="post" enctype="multipart/form-data">
      <input id="mdb" type="file" name="mdb" accept=".mdb" style="display:none" />
      <button type="button" onclick="document.getElementById('mdb').click()">Select MDB</button>
      <span id="mdb-name" class="file-name"></span>
      <br/><br/>
      <input id="excel" type="file" name="excel" accept=".xlsx" style="display:none" />
      <button type="button" onclick="document.getElementById('excel').click()">Select Excel</button>
      <span id="excel-name" class="file-name"></span>
      <br/><br/>
      <button type="submit">Upload</button>
    </form>
    <script>
      const mdbInput = document.getElementById('mdb');
      const mdbName = document.getElementById('mdb-name');
      mdbInput.addEventListener('change', () => {
        const file = mdbInput.files[0];
        mdbName.textContent = file ? file.name : '';
      });
      const excelInput = document.getElementById('excel');
      const excelName = document.getElementById('excel-name');
      excelInput.addEventListener('change', () => {
        const file = excelInput.files[0];
        excelName.textContent = file ? file.name : '';
      });
    </script>
    {% if message %}
    <p>{{ message }}</p>
    {% endif %}
    {% if mdb_available %}
    <p><a href="{{ url_for('download_updated_mdb') }}">Download Updated MDB</a></p>
    {% endif %}
    <p><a href="{{ url_for('home') }}">Back to Home</a></p>
  </div>
</body>
</html>
"""

PLC_TEMPLATE = """
<!doctype html>
<html lang='en'>
<head>
  <meta charset='utf-8'>
  <title>Read PLC Info</title>
  <style>
    body { font-family: Arial, sans-serif; background:#f2f2f2; margin:40px; }
    .container { max-width:600px; margin:auto; background:#fff; padding:40px; border-radius:8px; box-shadow:0 2px 4px rgba(0,0,0,0.1); }
    label { display:block; margin-top:10px; }
    input { padding:6px; margin-left:10px; }
    button { padding:8px 16px; margin-top:10px; }
  </style>
</head>
<body>
  <div class="container">
    <h1>Read PLC Info</h1>
    <form method="post">
      <label>IP Address: <input type="text" name="ip" value="{{ request.form.get('ip', '') }}"/></label>
      <label>Slot: <input type="text" name="slot" value="{{ request.form.get('slot', '0') }}"/></label>
      <button type="submit">Connect</button>
    </form>
    {% if message %}
    <p>{{ message }}</p>
    {% endif %}
    {% if info %}
    <h2>Information</h2>
    <ul>
      <li>Vendor: {{ info.vendor }}</li>
      <li>Product Type: {{ info.product_type }}</li>
      <li>Product Code: {{ info.product_code }}</li>
      <li>Revision: {{ info.revision.major }}.{{ info.revision.minor }}</li>
      <li>Serial: {{ info.serial }}</li>
      <li>Product Name: {{ info.product_name }}</li>
      <li>Keyswitch: {{ info.keyswitch }}</li>
      <li>Name: {{ info.name }}</li>
    </ul>
    <form method="post" action="{{ url_for('plc_tags') }}">
      <input type="hidden" name="ip" value="{{ request.form.get('ip', '') }}" />
      <input type="hidden" name="slot" value="{{ request.form.get('slot', '0') }}" />
      <button type="submit">Get Tag List</button>
    </form>
    {% endif %}
    {% if tags_available %}
    <p><a href="{{ url_for('download_tags') }}">Download Tag CSV</a></p>
    {% endif %}
    <p><a href="{{ url_for('home') }}">Back to Home</a></p>
  </div>
</body>
</html>
"""

app = Flask(__name__)
app.config['EXCEL_PATH'] = None
app.config['UPDATED_MDB_PATH'] = None
app.config['TAG_CSV_PATH'] = None

@app.route('/')
def home():
    """Landing page allowing navigation to export or import tools."""
    return render_template_string(HOME_TEMPLATE)


@app.route('/export', methods=['GET', 'POST'])
def export_page():
    message = None
    excel_available = False

    if request.method == 'POST':
        file = request.files.get('file')
        if file and file.filename:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.mdb') as tmp:
                file.save(tmp.name)
                mdb_path = tmp.name

            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as xlsx_tmp:
                xlsx_path = xlsx_tmp.name

            conn_str = (
                r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
                f'DBQ={mdb_path};'
            )

            try:
                conn = pyodbc.connect(conn_str, autocommit=True)
                cursor = conn.cursor()
                table_names = [row.table_name for row in cursor.tables(tableType='TABLE')]

                if 'Instruments' not in table_names:
                    message = 'Uploaded file does not contain an Instruments table.'
                else:
                    count = export_instruments_to_excel(mdb_path, xlsx_path)
                    app.config['EXCEL_PATH'] = xlsx_path
                    excel_available = True
                    message = f'MDB upload successful. Found {count} valid instruments.'

                conn.close()
            except (pyodbc.Error, ValueError) as e:
                message = f'Error accessing MDB file: {e}'
                try:
                    os.remove(xlsx_path)
                except Exception:
                    pass
            finally:
                os.remove(mdb_path)

    return render_template_string(
        EXPORT_TEMPLATE,
        message=message,
        excel_available=excel_available,
    )


@app.route('/download_excel')
def download_excel():
    """Send the exported instruments Excel file to the client."""
    xlsx_path = app.config.get('EXCEL_PATH')
    if not xlsx_path or not os.path.exists(xlsx_path):
        return "No Excel file available", 404

    @after_this_request
    def remove_file(response):
        try:
            os.remove(xlsx_path)
        except Exception:
            pass
        app.config['EXCEL_PATH'] = None
        return response

    return send_file(xlsx_path, as_attachment=True, download_name='instruments.xlsx')


@app.route('/import', methods=['GET', 'POST'])
def import_excel():
    message = None
    mdb_available = False

    if request.method == 'POST':
        mdb_file = request.files.get('mdb')
        excel_file = request.files.get('excel')
        if mdb_file and excel_file and mdb_file.filename and excel_file.filename:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.mdb') as mdb_tmp:
                mdb_file.save(mdb_tmp.name)
                mdb_path = mdb_tmp.name

            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as xls_tmp:
                excel_file.save(xls_tmp.name)
                xlsx_path = xls_tmp.name

            try:
                updated = update_instruments_from_excel(mdb_path, xlsx_path)
                app.config['UPDATED_MDB_PATH'] = mdb_path
                mdb_available = True
                message = f'MDB updated successfully. {updated} rows modified.'
            except Exception as e:
                message = f'Error updating MDB: {e}'
                os.remove(mdb_path)
            finally:
                os.remove(xlsx_path)
        else:
            message = 'Both MDB and Excel files are required.'

    return render_template_string(
        IMPORT_TEMPLATE,
        message=message,
        mdb_available=mdb_available,
    )


@app.route('/download_updated_mdb')
def download_updated_mdb():
    """Allow the user to download the updated MDB file."""
    mdb_path = app.config.get('UPDATED_MDB_PATH')
    if not mdb_path or not os.path.exists(mdb_path):
        return "No MDB file available", 404

    @after_this_request
    def remove_file(response):
        try:
            os.remove(mdb_path)
        except Exception:
            pass
        app.config['UPDATED_MDB_PATH'] = None
        return response

    return send_file(mdb_path, as_attachment=True, download_name='updated.mdb')


@app.route('/plc/tags', methods=['POST'])
def plc_tags():
    """Retrieve tag list from the PLC and store as CSV."""
    info = None
    message = None
    tags_available = False

    ip = request.form.get('ip', '').strip()
    slot = request.form.get('slot', '0').strip()
    if ip:
        path = f"{ip}/{slot}" if slot else ip
        try:
            with LogixDriver(path) as plc:
                plc.get_plc_info()
                info = {k: plc.info.get(k) for k in (
                    'vendor',
                    'product_type',
                    'product_code',
                    'revision',
                    'serial',
                    'product_name',
                    'keyswitch',
                    'name',
                )}

                tags = plc.get_tag_list()

            # open temporary CSV in text mode so csv module can write strings
            with tempfile.NamedTemporaryFile(delete=False, suffix='.csv', mode='w', newline='') as tmp:
                if tags:
                    fieldnames = sorted(tags[0].keys())
                    writer = csv.DictWriter(tmp, fieldnames=fieldnames)
                    writer.writeheader()
                    for t in tags:
                        writer.writerow({k: t.get(k, '') for k in fieldnames})
                else:
                    writer = csv.writer(tmp)
                    writer.writerow(['No tags found'])

                csv_path = tmp.name

            app.config['TAG_CSV_PATH'] = csv_path
            tags_available = True
            message = f'Tag list retrieved ({len(tags)} tags).'
        except Exception as exc:
            message = f'Error retrieving tag list: {exc}'
    else:
        message = 'IP address is required.'

    return render_template_string(
        PLC_TEMPLATE,
        info=info,
        message=message,
        tags_available=tags_available,
    )


@app.route('/download_tags')
def download_tags():
    """Allow the user to download the generated tag CSV file."""
    csv_path = app.config.get('TAG_CSV_PATH')
    if not csv_path or not os.path.exists(csv_path):
        return "No tag CSV available", 404

    @after_this_request
    def remove_file(response):
        try:
            os.remove(csv_path)
        except Exception:
            pass
        app.config['TAG_CSV_PATH'] = None
        return response

    return send_file(csv_path, as_attachment=True, download_name='tags.csv')


@app.route('/plc', methods=['GET', 'POST'])
def plc_page():
    """Connect to a PLC and display basic information."""
    info = None
    message = None
    tags_available = False

    if request.method == 'POST':
        ip = request.form.get('ip', '').strip()
        slot = request.form.get('slot', '0').strip()
        if ip:
            path = f"{ip}/{slot}" if slot else ip
            try:
                with LogixDriver(path) as plc:
                    plc.get_plc_info()
                    info = {k: plc.info.get(k) for k in (
                        'vendor',
                        'product_type',
                        'product_code',
                        'revision',
                        'serial',
                        'product_name',
                        'keyswitch',
                        'name',
                    )}
            except Exception as exc:
                message = f'Error connecting to PLC: {exc}'
        else:
            message = 'IP address is required.'

    return render_template_string(
        PLC_TEMPLATE,
        info=info,
        message=message,
        tags_available=tags_available,
    )

if __name__ == '__main__':
    app.run(debug=True)
